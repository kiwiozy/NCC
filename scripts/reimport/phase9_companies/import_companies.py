#!/usr/bin/env python3
"""
Import Companies from Excel (Companies.xlsx)

Reads company data from Excel file and creates Company records.

Usage:
    python3 import_companies.py --dry-run
    python3 import_companies.py
"""
import os
import sys
from pathlib import Path
from openpyxl import load_workbook

# Add backend to path
backend_dir = Path(__file__).parent.parent.parent.parent / 'backend'
sys.path.insert(0, str(backend_dir))

# Django setup
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ncc_api.settings')
import django
django.setup()

from django.db import transaction
from companies.models import Company

# Add project root for utils
project_root = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(project_root))
from scripts.reimport.utils.logger import create_logger
from scripts.reimport.utils.progress_tracker import ProgressTracker


def import_companies(excel_file: str = None, dry_run: bool = False) -> bool:
    logger = create_logger("PHASE 9")
    logger.phase_start("Phase 9", "Import Companies from Excel")

    if dry_run:
        logger.warning("🔍 DRY RUN MODE - No changes will be made")
        logger.info("")

    # Find Excel file
    if not excel_file:
        possible_paths = [
            Path.cwd() / 'Companies.xlsx',
            Path(__file__).parent.parent.parent.parent / 'Companies.xlsx',
            Path(__file__).parent.parent.parent.parent / 'scripts' / 'Export_Filemaker' / 'Companies.xlsx',
        ]

        for path in possible_paths:
            if path.exists():
                excel_file = str(path)
                break

        if not excel_file:
            logger.error("❌ Could not find Companies.xlsx")
            logger.error("Please provide --excel-file path")
            logger.phase_end(success=False)
            return False

    excel_path = Path(excel_file)
    if not excel_path.exists():
        logger.error(f"❌ Excel file not found: {excel_file}")
        logger.phase_end(success=False)
        return False

    logger.info(f"Reading companies from: {excel_file}")
    logger.info("Loading Excel workbook...")

    try:
        workbook = load_workbook(filename=excel_file, read_only=True, data_only=True)
        sheet = workbook.active

        # Get headers
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        logger.success(f"✅ Found {len(headers)} columns")

    except Exception as e:
        logger.error(f"Failed to read Excel file: {e}")
        logger.phase_end(success=False)
        return False

    # Process companies
    logger.info("")
    logger.info("Processing companies...")

    stats = {
        'total': 0,
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': 0,
    }

    companies_to_create = []
    companies_to_update = []
    row_count = 0

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
        if not any(row):
            continue

        row_data = dict(zip(headers, row))
        row_count += 1
        stats['total'] += 1

        try:
            # Extract fields
            filemaker_id = row_data.get('id')
            name = row_data.get('Name') or ''  # Fixed: Column is 'Name', not 'Company'
            abn = row_data.get('ABN') or ''
            company_type = row_data.get('Type') or 'OTHER'

            # Map company type
            type_mapping = {
                'Medical Practice': 'MEDICAL_PRACTICE',
                'NDIS Provider': 'NDIS_PROVIDER',
                'Other': 'OTHER',
            }
            company_type = type_mapping.get(company_type, 'OTHER')

            # Skip if no name
            if not name or not filemaker_id:
                stats['skipped'] += 1
                continue

            # Build contact JSON
            contact_json = {}
            if row_data.get('Phone'):
                contact_json['phone'] = str(row_data.get('Phone'))
            if row_data.get('Email'):
                contact_json['email'] = str(row_data.get('Email'))
            if row_data.get('Fax'):
                contact_json['fax'] = str(row_data.get('Fax'))

            # Build address JSON
            address_json = {}
            if row_data.get('Street'):
                address_json['street'] = str(row_data.get('Street'))
            if row_data.get('Street2'):
                address_json['street2'] = str(row_data.get('Street2'))
            if row_data.get('Suburb'):
                address_json['suburb'] = str(row_data.get('Suburb'))
            if row_data.get('State'):
                address_json['state'] = str(row_data.get('State'))
            if row_data.get('Postcode'):
                address_json['postcode'] = str(row_data.get('Postcode'))

            # Check if company exists
            existing_company = None
            if filemaker_id:
                try:
                    existing_company = Company.objects.get(filemaker_id=filemaker_id)
                except Company.DoesNotExist:
                    pass

            if existing_company:
                # Update existing
                existing_company.name = name
                existing_company.abn = abn
                existing_company.company_type = company_type
                existing_company.contact_json = contact_json or {}
                existing_company.address_json = address_json or {}
                companies_to_update.append(existing_company)
                stats['updated'] += 1
            else:
                # Create new
                company = Company(
                    name=name,
                    abn=abn,
                    company_type=company_type,
                    contact_json=contact_json or {},
                    address_json=address_json or {},
                    filemaker_id=filemaker_id,
                )
                companies_to_create.append(company)
                stats['created'] += 1

        except Exception as e:
            logger.error(f"Error processing row {row_count}: {e}")
            stats['errors'] += 1

        # Progress every 10 rows
        if row_count % 10 == 0:
            logger.info(f"💓 Processed {row_count:,} rows...")

    logger.success(f"✅ Processed {row_count:,} rows")

    # Save to database
    if not dry_run:
        logger.info("")
        logger.info("Saving to database...")

        with transaction.atomic():
            if companies_to_create:
                Company.objects.bulk_create(companies_to_create, batch_size=100)
                logger.success(f"✅ Created {len(companies_to_create):,} companies")

            if companies_to_update:
                Company.objects.bulk_update(
                    companies_to_update,
                    ['name', 'abn', 'company_type', 'contact_json', 'address_json'],
                    batch_size=100
                )
                logger.success(f"✅ Updated {len(companies_to_update):,} companies")

    # Summary
    logger.info("")
    logger.info("=" * 70)
    logger.info("📊 Companies Import Summary")
    logger.info("=" * 70)
    logger.info(f"Total Rows: {stats['total']:,}")
    logger.success(f"✅ Created: {stats['created']:,}")
    logger.success(f"✅ Updated: {stats['updated']:,}")
    logger.info(f"⏭️  Skipped: {stats['skipped']:,}")
    logger.error(f"❌ Errors: {stats['errors']:,}")
    logger.info("")

    if stats['errors'] == 0:
        logger.success("✅ ✅ ✅ Companies import completed successfully!")
        success = True
    else:
        logger.error("❌ Companies import completed with errors!")
        success = False

    logger.phase_end(success=success)
    return success


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Import companies from Excel')
    parser.add_argument('--excel-file', help='Path to Companies.xlsx')
    parser.add_argument('--dry-run', action='store_true', help='Dry run mode')
    args = parser.parse_args()

    success = import_companies(excel_file=args.excel_file, dry_run=args.dry_run)
    sys.exit(0 if success else 1)

