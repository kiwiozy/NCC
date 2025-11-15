#!/usr/bin/env python3
"""
Import Patients from Contacts.xlsx (Excel-Based Import)

This script reads patient data from Excel export instead of FileMaker API.
Much faster and more reliable than API-based import.

Usage:
    python import_patients_from_excel.py
    python import_patients_from_excel.py --dry-run
    python import_patients_from_excel.py --excel-file ../Export_Filemaker/Contacts.xlsx
"""
import sys
import os
from pathlib import Path
from datetime import datetime

# Add project paths
project_root = Path(__file__).parent.parent.parent.parent
backend_dir = project_root / 'backend'
sys.path.insert(0, str(backend_dir))
sys.path.insert(0, str(project_root))

# Django setup
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ncc_api.settings')
import django
django.setup()

from patients.models import Patient
from clinicians.models import Clinic
from settings.models import FundingSource
from django.db import transaction
from scripts.reimport.utils.logger import create_logger

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed")
    print("Install with: pip install openpyxl")
    sys.exit(1)


def parse_date(date_value):
    """
    Parse date from various formats.
    
    Args:
        date_value: Date string or datetime object
    
    Returns:
        Date object or None
    """
    if not date_value:
        return None
    
    # If already a datetime object
    if isinstance(date_value, datetime):
        return date_value.date()
    
    # Try string parsing
    date_str = str(date_value).strip()
    if not date_str or date_str.lower() == 'none':
        return None
    
    # Try various formats
    for fmt in ['%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
        try:
            return datetime.strptime(date_str, fmt).date()
        except:
            continue
    
    return None


def transform_gender(gender_str):
    """
    Transform gender to standard format.
    
    Args:
        gender_str: Gender string from FileMaker
    
    Returns:
        Standardized gender ('M', 'F', 'O', or empty)
    """
    if not gender_str:
        return ''
    
    gender = str(gender_str).strip().upper()
    
    if gender in ['M', 'MALE', 'MAN']:
        return 'M'
    elif gender in ['F', 'FEMALE', 'WOMAN']:
        return 'F'
    elif gender in ['O', 'OTHER', 'NON-BINARY', 'X']:
        return 'O'
    else:
        return ''


def import_patients_from_excel(excel_file: str = None, dry_run: bool = False) -> bool:
    """
    Import patients from Contacts.xlsx.
    
    Args:
        excel_file: Path to Excel file (default: scripts/Export_Filemaker/Contacts.xlsx)
        dry_run: If True, preview changes without modifying database
    
    Returns:
        True if successful, False otherwise
    """
    logger = create_logger("PATIENTS")
    logger.phase_start("Phase 3", "Import Patients from Excel")
    
    if dry_run:
        logger.warning("🔍 DRY RUN MODE - No changes will be made")
        logger.info("")
    
    # ========================================
    # Determine Excel File Path
    # ========================================
    if excel_file is None:
        excel_file = project_root / 'scripts' / 'Export_Filemaker' / 'Contacts.xlsx'
    else:
        excel_file = Path(excel_file)
        if not excel_file.is_absolute():
            # Check project root if relative path
            root_path = project_root / excel_file
            if root_path.exists():
                excel_file = root_path
    
    if not excel_file.exists():
        logger.error(f"Excel file not found: {excel_file}")
        logger.phase_end(success=False)
        return False
    
    logger.info(f"📊 Reading patients from: {excel_file.name}")
    logger.info("")
    
    # ========================================
    # Load Excel File
    # ========================================
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active
        
        # Get headers
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value))
        
        # Count total rows
        total_rows = 0
        for row in sheet.iter_rows(min_row=2):
            if any(cell.value for cell in row):
                total_rows += 1
        
        logger.success(f"✅ Loaded {total_rows:,} patient records")
        logger.info(f"   Columns: {len(headers)}")
    except Exception as e:
        logger.error(f"Failed to load Excel file: {e}")
        logger.phase_end(success=False)
        return False
    
    # ========================================
    # Load Nexus Configuration
    # ========================================
    logger.info("")
    logger.info("=" * 70)
    logger.info("Step 1: Loading Nexus configuration...")
    logger.info("=" * 70)
    
    try:
        clinics = {clinic.name: clinic for clinic in Clinic.objects.all()}
        funding_sources = {fs.name: fs for fs in FundingSource.objects.all()}
        
        logger.success(f"✅ Loaded {len(clinics)} clinics")
        logger.success(f"✅ Loaded {len(funding_sources)} funding sources")
    except Exception as e:
        logger.error(f"Failed to load configuration: {e}")
        logger.phase_end(success=False)
        return False
    
    # ========================================
    # Import Patients
    # ========================================
    logger.info("")
    logger.info("=" * 70)
    logger.info("Step 2: Importing patients...")
    logger.info("=" * 70)
    logger.info("")
    
    stats = {
        'total': 0,
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': 0,
        'with_funding': 0,
        'without_funding': 0
    }
    
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
        if not any(row):
            continue
        
        row_data = dict(zip(headers, row))
        stats['total'] += 1
        
        # Progress heartbeat every 50 records
        if stats['total'] % 50 == 0:
            logger.info(f"  💓 Still working... {stats['total']:,}/{total_rows:,} patients processed")
        
        # Progress bar every 100 records
        if stats['total'] % 100 == 0:
            logger.progress(stats['total'], total_rows, "Importing patients")
        
        try:
            # ========================================
            # Extract Patient Data
            # ========================================
            filemaker_id = row_data.get('id')
            first_name = (row_data.get('nameFirst') or '').strip()
            last_name = (row_data.get('nameLast') or '').strip()
            middle_name = (row_data.get('nameMiddle') or '').strip()
            title = (row_data.get('title') or '').strip()
            
            # Skip if no name
            if not last_name:
                stats['skipped'] += 1
                continue
            
            # Parse DOB
            dob = parse_date(row_data.get('DOB'))
            
            # Transform gender
            gender = transform_gender(row_data.get('gender'))
            
            # Get clinic
            clinic_name = row_data.get('Clinic_Name')
            clinic = clinics.get(clinic_name) if clinic_name else None
            
            # Get funding source
            funding_name = row_data.get('Funding')
            funding_source = funding_sources.get(funding_name) if funding_name else None
            
            if funding_source:
                stats['with_funding'] += 1
            else:
                stats['without_funding'] += 1
            
            # ========================================
            # Build Patient Data
            # ========================================
            patient_data = {
                'first_name': first_name,
                'last_name': last_name,
                'middle_names': middle_name or '',  # Note: field is 'middle_names' not 'middle_name'
                'title': title or '',
                'dob': dob,  # Note: field is 'dob' not 'date_of_birth'
                'sex': gender,  # Note: field is 'sex' not 'gender'
                'health_number': row_data.get('Health Number') or '',  # Health number (Medicare, etc.)
                'clinic': clinic,
                'funding_type': funding_source,  # Note: field is 'funding_type' not 'funding_source'
                'filemaker_metadata': {
                    'filemaker_id': str(filemaker_id),
                    'clinic_name': clinic_name or '',
                    'area': row_data.get('Area') or '',
                    'contact_type': row_data.get('contactType') or '',
                    'cys_id': row_data.get('CYS_ID') or '',
                    'diabetes': row_data.get('Diabetes') or '',
                    'id_clinic': row_data.get('id_Clinic') or '',
                },
                'notes': ''  # Will be populated by Phase 5
            }
            
            # Add NDIS data if present
            ndis_coordinator_name = row_data.get('NDIS Coordinator Name')
            if ndis_coordinator_name:
                patient_data['filemaker_metadata']['ndis'] = {
                    'coordinator_name': ndis_coordinator_name,
                    'coordinator_email': row_data.get('NDIS Coordinator Email') or '',
                    'coordinator_phone': row_data.get('NDIS Coordinator Phone') or '',
                    'plan_start_date': str(row_data.get('NDIS Plan Start Date') or ''),
                    'plan_end_date': str(row_data.get('NDIS Plan End Date') or ''),
                    'notes': row_data.get('NDIS notes') or '',
                }
            
            # ========================================
            # Save Patient
            # ========================================
            if not dry_run:
                try:
                    with transaction.atomic():
                        # Check if patient exists by FileMaker ID
                        existing = Patient.objects.filter(
                            filemaker_metadata__filemaker_id=str(filemaker_id)
                        ).first()
                        
                        if existing:
                            # Update existing patient
                            for key, value in patient_data.items():
                                setattr(existing, key, value)
                            existing.save()
                            stats['updated'] += 1
                        else:
                            # Create new patient
                            Patient.objects.create(**patient_data)
                            stats['created'] += 1
                except Exception as e:
                    logger.error(f"Error importing patient {filemaker_id}: {e}")
                    stats['errors'] += 1
            else:
                # Dry run - just count as created
                stats['created'] += 1
        
        except Exception as e:
            logger.error(f"Error processing row {stats['total']}: {e}")
            stats['errors'] += 1
    
    # ========================================
    # Summary
    # ========================================
    logger.info("")
    logger.info("=" * 70)
    logger.info("📊 SUMMARY")
    logger.info("=" * 70)
    logger.info(f"Total processed:       {stats['total']:,}")
    logger.info(f"Created:               {stats['created']:,}")
    logger.info(f"Updated:               {stats['updated']:,}")
    logger.info(f"Skipped:               {stats['skipped']:,}")
    logger.info(f"Errors:                {stats['errors']:,}")
    logger.info("")
    logger.info(f"With funding source:   {stats['with_funding']:,}")
    logger.info(f"Without funding:       {stats['without_funding']:,}")
    logger.info("")
    
    if dry_run:
        logger.warning("🔍 DRY RUN - No changes made to database")
    else:
        logger.success("✅ Patient import complete!")
    
    logger.info("")
    wb.close()
    logger.phase_end(success=True)
    return True


def main():
    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Import patients from Contacts.xlsx'
    )
    parser.add_argument(
        '--excel-file',
        default=None,
        help='Path to Contacts Excel file (default: scripts/Export_Filemaker/Contacts.xlsx)'
    )
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Preview changes without modifying database'
    )
    
    args = parser.parse_args()
    
    success = import_patients_from_excel(
        excel_file=args.excel_file,
        dry_run=args.dry_run
    )
    
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()

