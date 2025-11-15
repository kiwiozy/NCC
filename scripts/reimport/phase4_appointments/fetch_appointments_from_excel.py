"""
Phase 4: Fetch Appointments from Excel Export

Reads appointments from Appointments.xlsx file exported from FileMaker.
Exports to JSON file for import into Nexus.
"""

import sys
import os
import json
from datetime import datetime
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from utils import create_logger


def fetch_appointments_from_excel(excel_file: str = "Appointments.xlsx", output_dir: str = "data/reimport") -> str:
    """
    Fetch all appointments from Excel file.
    
    Args:
        excel_file: Path to Appointments.xlsx file
        output_dir: Directory to save export file
    
    Returns:
        Path to exported JSON file, or None if failed
    """
    logger = create_logger("PHASE 4")
    logger.phase_start("Phase 4.1", "Fetch Appointments from Excel")
    
    try:
        from openpyxl import load_workbook
        
        # Create output directory
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        # Create timestamped filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        appointments_file = output_path / f"appointments_export_{timestamp}.json"
        
        logger.info(f"Appointments will be saved to: {appointments_file}")
        
        # ========================================
        # Read Appointments from Excel
        # ========================================
        # If relative path, look in project root
        excel_path = Path(excel_file)
        if not excel_path.is_absolute():
            # Look in project root (4 levels up from this script)
            project_root = Path(__file__).parent.parent.parent.parent
            excel_path = project_root / excel_file
        if not excel_path.exists():
            logger.error(f"Excel file not found: {excel_file}")
            logger.phase_end(success=False)
            return None
        
        logger.info(f"Reading appointments from: {excel_file}")
        
        wb = load_workbook(excel_path)
        sheet = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]
        logger.info(f"Excel columns: {', '.join([h for h in headers if h])}")
        
        # Read all rows
        all_appointments = []
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            if (i + 1) % 1000 == 0:
                logger.info(f"💓 Processed {i + 1} appointments...")
            
            row_data = dict(zip(headers, row))
            
            # Extract key fields
            appointment_id = row_data.get('id')
            if not appointment_id:
                logger.warning(f"Skipping row {i+2}: No 'id' found.")
                continue
            
            # Transform data
            transformed_appointment = {
                'id': str(appointment_id),
                'id_Contact': str(row_data.get('id_Contact')) if row_data.get('id_Contact') else None,
                'id_Clinic': str(row_data.get('id_Clinic')) if row_data.get('id_Clinic') else None,
                'id_Clinician': str(row_data.get('id_Clinician')) if row_data.get('id_Clinician') else None,
                'startDate': row_data.get('startDate').isoformat() if isinstance(row_data.get('startDate'), datetime) else row_data.get('startDate'),
                'startTime': row_data.get('startTime'),
                'endDate': row_data.get('endDate').isoformat() if isinstance(row_data.get('endDate'), datetime) else row_data.get('endDate'),
                'endTime': row_data.get('endTime'),
                'Type': row_data.get('Type'),
                'notes': row_data.get('notes'),
                'source': 'FileMaker Excel Export'
            }
            
            all_appointments.append(transformed_appointment)
        
        logger.success(f"✅ Read {len(all_appointments)} appointments from Excel")
        
        # Analyze appointment data
        appointments_with_clinic = sum(1 for a in all_appointments if a.get('id_Clinic'))
        appointments_with_clinician = sum(1 for a in all_appointments if a.get('id_Clinician'))
        appointments_with_type = sum(1 for a in all_appointments if a.get('Type'))
        appointments_with_patient = sum(1 for a in all_appointments if a.get('id_Contact'))
        appointments_with_start = sum(1 for a in all_appointments if a.get('startDate'))
        
        logger.info(f"Appointments with clinic: {appointments_with_clinic}/{len(all_appointments)}")
        logger.info(f"Appointments with clinician: {appointments_with_clinician}/{len(all_appointments)}")
        logger.info(f"Appointments with type: {appointments_with_type}/{len(all_appointments)}")
        logger.info(f"Appointments with patient: {appointments_with_patient}/{len(all_appointments)}")
        logger.info(f"Appointments with start date: {appointments_with_start}/{len(all_appointments)}")
        
        # Save appointments to JSON
        logger.info("Writing JSON export...")
        appointments_export = {
            'export_timestamp': timestamp,
            'export_date': datetime.now().isoformat(),
            'total_appointments': len(all_appointments),
            'source': 'FileMaker Excel Export (Appointments.xlsx)',
            'appointments': all_appointments
        }
        
        with open(appointments_file, 'w') as f:
            json.dump(appointments_export, f, indent=2, default=str)
        
        logger.success(f"✅ Saved appointments to: {appointments_file} ({appointments_file.stat().st_size / (1024*1024):.2f} MB)")
        
        logger.info("")
        logger.info("=" * 70)
        logger.info("📊 Export Summary")
        logger.info("=" * 70)
        logger.info(f"Total Appointments: {len(all_appointments)}")
        logger.info(f"Appointments File: {appointments_file} ({appointments_file.stat().st_size / (1024*1024):.2f} MB)")
        logger.info("")
        logger.success("✅ ✅ ✅ Appointment export from Excel completed successfully!")
        logger.success("✅ ✅ Next: Run import_appointments.py to import into Nexus")
        logger.success("")
        
        logger.phase_end(success=True)
        return str(appointments_file)
        
    except Exception as e:
        logger.error(f"❌ Exception reading Excel file: {e}", exc_info=e)
        logger.phase_end(success=False)
        return None


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Fetch appointments from Excel file.')
    parser.add_argument('--excel-file', default='Appointments.xlsx', help='Path to the Appointments Excel file.')
    parser.add_argument('--output-dir', default='data/reimport', help='Output directory for export files.')
    args = parser.parse_args()
    
    result = fetch_appointments_from_excel(excel_file=args.excel_file, output_dir=args.output_dir)
    sys.exit(0 if result else 1)

