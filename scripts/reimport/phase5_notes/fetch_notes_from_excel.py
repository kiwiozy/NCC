"""
Phase 5: Fetch Notes from Excel Export

Reads notes from Notes.xlsx file exported from FileMaker.
Exports to JSON file for import into Nexus.
"""

import sys
import os
import json
from datetime import datetime
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from utils import create_logger


def fetch_notes_from_excel(excel_file: str = "Notes.xlsx", output_dir: str = "data/reimport") -> str:
    """
    Fetch all notes from Excel file.
    
    Args:
        excel_file: Path to Notes.xlsx file
        output_dir: Directory to save export file
    
    Returns:
        Path to exported JSON file, or None if failed
    """
    logger = create_logger("PHASE 5")
    logger.phase_start("Phase 5.1", "Fetch Notes from Excel")
    
    try:
        from openpyxl import load_workbook
        
        # Create output directory
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        # Create timestamped filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        notes_file = output_path / f"notes_export_{timestamp}.json"
        
        logger.info(f"Notes will be saved to: {notes_file}")
        
        # ========================================
        # Read Notes from Excel
        # ========================================
        excel_path = Path(excel_file)
        if not excel_path.exists():
            # Try in project root
            excel_path = Path(__file__).parent.parent.parent.parent / excel_file
        
        if not excel_path.exists():
            logger.error(f"Excel file not found: {excel_file}")
            logger.error(f"Tried: {excel_path}")
            logger.phase_end(success=False)
            return None
        
        logger.info(f"Reading notes from: {excel_path}")
        logger.info("Loading Excel workbook...")
        
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        logger.success(f"✅ Found {len(headers)} columns")
        
        # Read all rows
        logger.info("Reading notes...")
        notes = []
        row_count = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue
            
            # Create note dict from row
            note = {}
            for i, header in enumerate(headers):
                value = row[i] if i < len(row) else None
                
                # Convert datetime to string
                if isinstance(value, datetime):
                    value = value.isoformat()
                
                note[header] = value
            
            notes.append(note)
            row_count += 1
            
            # Progress update every 1000 rows
            if row_count % 1000 == 0:
                logger.info(f"💓 Processed {row_count} notes...")
        
        wb.close()
        
        logger.success(f"✅ Read {len(notes)} notes from Excel")
        
        # Analyze note data
        if notes:
            notes_with_patient = sum(1 for n in notes if n.get('@Notes to Contacts::Name Full'))
            notes_with_content = sum(1 for n in notes if n.get('Note'))
            notes_with_date = sum(1 for n in notes if n.get('Date'))
            
            logger.info(f"Notes with patient link: {notes_with_patient}/{len(notes)}")
            logger.info(f"Notes with content: {notes_with_content}/{len(notes)}")
            logger.info(f"Notes with date: {notes_with_date}/{len(notes)}")
        
        # Save notes
        if notes:
            notes_export = {
                'export_timestamp': timestamp,
                'export_date': datetime.now().isoformat(),
                'total_notes': len(notes),
                'source': f'Excel export: {excel_path.name}',
                'notes': notes
            }
            
            logger.info("Writing JSON export...")
            with open(notes_file, 'w') as f:
                json.dump(notes_export, f, indent=2, default=str)
            
            file_size_mb = notes_file.stat().st_size / (1024 * 1024)
            logger.success(f"✅ Saved notes to: {notes_file} ({file_size_mb:.2f} MB)")
        
        # ========================================
        # Export Summary
        # ========================================
        logger.info("")
        logger.info("=" * 70)
        logger.info("📊 Export Summary")
        logger.info("=" * 70)
        logger.info(f"Total Notes: {len(notes)}")
        logger.info(f"Notes File: {notes_file} ({file_size_mb:.2f} MB)")
        logger.info("")
        logger.success("✅ ✅ Note export from Excel completed successfully!")
        logger.success("✅ Next: Run import_notes.py to import into Nexus")
        logger.info("")
        
        logger.phase_end(success=True)
        return str(notes_file)
        
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        logger.phase_end(success=False)
        return None
    except Exception as e:
        logger.error(f"Exception during fetch: {str(e)}", exc_info=e)
        logger.phase_end(success=False)
        return None


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Fetch notes from Excel export')
    parser.add_argument('--excel', default='Notes.xlsx', help='Path to Notes.xlsx file')
    parser.add_argument('--output-dir', default='data/reimport', help='Output directory for export files')
    args = parser.parse_args()
    
    notes_file = fetch_notes_from_excel(excel_file=args.excel, output_dir=args.output_dir)
    sys.exit(0 if notes_file else 1)

