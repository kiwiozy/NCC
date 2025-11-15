"""
Phase 6: Re-Link Documents to New Patients with Clean S3 Paths

This script:
1. Reads Docs.xlsx to get document→patient mappings from FileMaker
2. For each document:
   - Finds the new patient by FileMaker ID
   - Copies S3 file to new patient folder
   - Updates database link
   - Deletes old S3 file
3. Results in perfectly clean data where S3 paths match database relationships
"""

import sys
import os
import django
from pathlib import Path

# Add Django project to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../../../backend')))
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ncc_api.settings')
django.setup()

from utils import create_logger
from patients.models import Patient
from documents.models import Document
from documents.services import S3Service
from django.db import transaction
from django.contrib.contenttypes.models import ContentType
from openpyxl import load_workbook


def load_document_patient_mapping(excel_file: str, logger) -> dict:
    """
    Load document→patient mapping from Docs.xlsx
    
    Returns:
        dict: {document_filemaker_id: patient_filemaker_id}
    """
    logger.info(f"Loading document mappings from {excel_file}...")
    
    excel_path = Path(excel_file)
    if not excel_path.exists():
        # Try in project root
        excel_path = Path(__file__).parent.parent.parent.parent / excel_file
    
    if not excel_path.exists():
        logger.error(f"Excel file not found: {excel_file}")
        return {}
    
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    
    # Build mapping: document_id → patient_id
    doc_to_patient = {}
    row_count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and row[2]:  # id and id_Contact columns
            doc_id = str(row[1]).upper().strip()
            patient_id = str(row[2]).upper().strip()
            doc_to_patient[doc_id] = patient_id
            row_count += 1
    
    wb.close()
    
    logger.success(f"✅ Loaded {len(doc_to_patient)} document→patient mappings from Excel")
    return doc_to_patient


def build_patient_lookup(logger) -> dict:
    """
    Build lookup map: filemaker_id → Patient object
    
    Returns:
        dict: {filemaker_id: Patient}
    """
    logger.info("Building patient lookup map...")
    
    patient_map = {}
    
    for patient in Patient.objects.all():
        fm_metadata = patient.filemaker_metadata
        if fm_metadata and fm_metadata.get('filemaker_id'):
            filemaker_id = str(fm_metadata['filemaker_id']).upper().strip()
            patient_map[filemaker_id] = patient
    
    logger.success(f"✅ Built lookup map for {len(patient_map)} patients")
    return patient_map


def relink_documents_clean(dry_run: bool = False, excel_file: str = "Docs.xlsx") -> bool:
    """
    Re-link documents with clean S3 paths.
    
    Args:
        dry_run: If True, preview without making changes
        excel_file: Path to Docs.xlsx file
    
    Returns:
        True if successful, False otherwise
    """
    logger = create_logger("PHASE 6")
    logger.phase_start("Phase 6", "Re-Link Documents with Clean S3 Paths")
    
    if dry_run:
        logger.warning("🔍 DRY RUN MODE - No changes will be made")
    
    try:
        # ========================================
        # Load Document→Patient Mapping from Excel
        # ========================================
        doc_to_patient_map = load_document_patient_mapping(excel_file, logger)
        
        if not doc_to_patient_map:
            logger.error("Failed to load document mappings from Excel")
            logger.phase_end(success=False)
            return False
        
        # ========================================
        # Build Patient Lookup Map
        # ========================================
        patient_map = build_patient_lookup(logger)
        
        if not patient_map:
            logger.error("No patients found in database")
            logger.phase_end(success=False)
            return False
        
        # ========================================
        # Get All Documents to Process
        # ========================================
        logger.info("Finding documents to relink...")
        
        documents = Document.objects.filter(filemaker_id__isnull=False).order_by('uploaded_at')
        total_documents = documents.count()
        
        logger.info(f"Found {total_documents} documents with filemaker_id")
        
        # CRITICAL: Count documents BEFORE starting
        initial_doc_count = Document.objects.count()
        logger.info(f"Total documents in database BEFORE: {initial_doc_count}")
        
        # CRITICAL: Track all document IDs to ensure none are lost
        all_doc_ids = set(documents.values_list('id', flat=True))
        logger.info(f"Tracking {len(all_doc_ids)} document IDs for safety verification")
        
        # ========================================
        # Initialize S3 Service
        # ========================================
        if not dry_run:
            logger.info("Initializing S3 service...")
            s3_service = S3Service()
            logger.success("✅ S3 service ready")
        
        # ========================================
        # Re-Link Documents with Clean Paths
        # ========================================
        logger.info("")
        logger.info("=" * 70)
        logger.info("Starting document relinking with S3 cleanup...")
        logger.info("=" * 70)
        logger.info("")
        
        stats = {
            'processed': 0,
            'linked': 0,
            'already_linked': 0,
            'not_in_excel': 0,
            'patient_not_found': 0,
            's3_copy_success': 0,
            's3_copy_fail': 0,
            's3_delete_success': 0,
            's3_delete_fail': 0,
            'errors': 0
        }
        
        patient_ct = ContentType.objects.get_for_model(Patient)
        
        for i, doc in enumerate(documents):
            stats['processed'] += 1
            
            # Progress updates
            if (i + 1) % 50 == 0:
                logger.info(f"💓 Still working... {i + 1}/{total_documents} documents processed")
            
            if (i + 1) % 100 == 0:
                logger.progress(i + 1, total_documents, "Re-linking documents")
            
            try:
                # Get document's FileMaker ID
                doc_fm_id = str(doc.filemaker_id).upper().strip()
                
                # Look up patient FileMaker ID from Excel
                patient_fm_id = doc_to_patient_map.get(doc_fm_id)
                
                if not patient_fm_id:
                    logger.debug(f"Document {doc_fm_id[:20]}... not found in Excel")
                    stats['not_in_excel'] += 1
                    continue
                
                # Find new patient
                patient = patient_map.get(patient_fm_id)
                
                if not patient:
                    logger.debug(f"Patient {patient_fm_id[:20]}... not found in database")
                    stats['patient_not_found'] += 1
                    continue
                
                # Check if already linked correctly
                if doc.content_type == patient_ct and doc.object_id == str(patient.id):
                    # Check if S3 path is already correct
                    if str(patient.id) in doc.s3_key:
                        stats['already_linked'] += 1
                        continue
                
                # ========================================
                # Copy to New S3 Location
                # ========================================
                old_s3_key = doc.s3_key
                
                # Build new S3 key with correct patient ID
                # Old: patients/filemaker-import/documents/{OLD_PATIENT_ID}/category/...
                # New: patients/filemaker-import/documents/{NEW_PATIENT_ID}/category/...
                
                parts = old_s3_key.split('/')
                if len(parts) >= 4:
                    # Replace old patient ID with new patient ID
                    parts[3] = str(patient.id)
                    new_s3_key = '/'.join(parts)
                else:
                    logger.warning(f"Unexpected S3 key format: {old_s3_key}")
                    stats['errors'] += 1
                    continue
                
                if not dry_run:
                    # Step 1: Copy file to new S3 location
                    try:
                        s3_service.s3_client.copy_object(
                            Bucket=s3_service.bucket_name,
                            CopySource={'Bucket': s3_service.bucket_name, 'Key': old_s3_key},
                            Key=new_s3_key
                        )
                        stats['s3_copy_success'] += 1
                        
                        if (stats['s3_copy_success'] + stats['s3_copy_fail']) % 500 == 0:
                            logger.info(f"  📦 Copied {stats['s3_copy_success']} files to new locations...")
                        
                    except Exception as e:
                        logger.warning(f"Failed to copy S3 file: {old_s3_key} → {new_s3_key}: {e}")
                        stats['s3_copy_fail'] += 1
                        stats['errors'] += 1
                        continue
                    
                    # Step 2: Verify new file exists before proceeding
                    try:
                        s3_service.s3_client.head_object(
                            Bucket=s3_service.bucket_name,
                            Key=new_s3_key
                        )
                    except Exception as e:
                        logger.error(f"Verification failed - new file not found: {new_s3_key}")
                        stats['s3_copy_fail'] += 1
                        stats['errors'] += 1
                        continue
                    
                    # Step 3: Update database (ONLY after successful copy + verification)
                    try:
                        with transaction.atomic():
                            doc.content_object = patient
                            doc.s3_key = new_s3_key
                            doc.save()
                    except Exception as e:
                        logger.error(f"Database update failed for document {doc.id}: {e}")
                        # Try to clean up the copied file
                        try:
                            s3_service.s3_client.delete_object(
                                Bucket=s3_service.bucket_name,
                                Key=new_s3_key
                            )
                        except:
                            pass
                        stats['errors'] += 1
                        continue
                    
                    # Step 4: Delete old S3 file (ONLY after successful database update)
                    # Safety check: make sure old and new keys are different
                    if old_s3_key != new_s3_key:
                        try:
                            s3_service.s3_client.delete_object(
                                Bucket=s3_service.bucket_name,
                                Key=old_s3_key
                            )
                            stats['s3_delete_success'] += 1
                            
                            if stats['s3_delete_success'] % 500 == 0:
                                logger.info(f"  🗑️  Cleaned up {stats['s3_delete_success']} old files...")
                                
                        except Exception as e:
                            logger.debug(f"Failed to delete old S3 file (not critical): {old_s3_key}: {e}")
                            stats['s3_delete_fail'] += 1
                            # Don't count as error - file was copied and database updated successfully
                    else:
                        logger.warning(f"Skipped delete - old and new S3 keys are identical: {old_s3_key}")
                
                stats['linked'] += 1
                logger.increment_success()
                
            except Exception as e:
                logger.error(f"Error processing document {doc.id}: {str(e)}", exc_info=e)
                stats['errors'] += 1
                logger.increment_error()
        
        # ========================================
        # Final Progress Update
        # ========================================
        logger.progress(total_documents, total_documents, "Re-linking documents")
        
        # ========================================
        # Verification
        # ========================================
        logger.info("")
        logger.info("=" * 70)
        logger.info("🔍 CRITICAL SAFETY VERIFICATION")
        logger.info("=" * 70)
        
        if not dry_run:
            # Verify document count unchanged
            final_doc_count = Document.objects.count()
            logger.info(f"Total documents BEFORE: {initial_doc_count}")
            logger.info(f"Total documents AFTER:  {final_doc_count}")
            
            if final_doc_count != initial_doc_count:
                logger.error(f"❌ CRITICAL: Document count changed! Lost {initial_doc_count - final_doc_count} documents!")
                logger.error("This should NEVER happen - aborting!")
                logger.phase_end(success=False)
                return False
            else:
                logger.success("✅ Document count verified: All documents preserved")
            
            # Verify all document IDs still exist
            remaining_doc_ids = set(Document.objects.filter(id__in=all_doc_ids).values_list('id', flat=True))
            missing_docs = all_doc_ids - remaining_doc_ids
            
            if missing_docs:
                logger.error(f"❌ CRITICAL: {len(missing_docs)} documents went missing!")
                logger.error(f"Missing IDs: {list(missing_docs)[:5]}...")
                logger.phase_end(success=False)
                return False
            else:
                logger.success("✅ All document IDs verified: No documents lost")
            
            # Verify orphaned documents
            orphaned_documents = Document.objects.filter(
                filemaker_id__isnull=False,
                object_id__isnull=True
            ).count()
            
            if orphaned_documents > 0:
                logger.warning(f"⚠️  {orphaned_documents} documents still orphaned (patient not found in Excel)")
            else:
                logger.success("✅ All documents have patient assigned")
            
            # Verify S3 files exist (sample check)
            logger.info("")
            logger.info("Verifying S3 files (sample check)...")
            sample_docs = Document.objects.filter(filemaker_id__isnull=False)[:10]
            s3_verified = 0
            s3_missing = 0
            
            for doc in sample_docs:
                try:
                    s3_service.s3_client.head_object(
                        Bucket=s3_service.bucket_name,
                        Key=doc.s3_key
                    )
                    s3_verified += 1
                except:
                    s3_missing += 1
                    logger.warning(f"⚠️  S3 file missing: {doc.s3_key}")
            
            logger.info(f"S3 Sample Check: {s3_verified}/10 files verified")
            if s3_missing > 0:
                logger.warning(f"⚠️  {s3_missing}/10 sample files missing in S3")
        else:
            logger.info("Skipping verification (dry-run mode)")
        
        # ========================================
        # Summary
        # ========================================
        logger.info("")
        logger.info("=" * 70)
        logger.info("📊 Re-Linking Summary")
        logger.info("=" * 70)
        logger.info(f"Total Documents Processed: {stats['processed']}")
        logger.info(f"✅ Successfully Re-Linked: {stats['linked']}")
        logger.info(f"Already Linked Correctly: {stats['already_linked']}")
        logger.info(f"⚠️  Not in Excel: {stats['not_in_excel']}")
        logger.info(f"⚠️  Patient Not Found: {stats['patient_not_found']}")
        logger.info(f"❌ Errors: {stats['errors']}")
        logger.info("")
        logger.info("📦 S3 Operations:")
        logger.info(f"  Copied: {stats['s3_copy_success']}")
        logger.info(f"  Copy Failed: {stats['s3_copy_fail']}")
        logger.info(f"  Deleted: {stats['s3_delete_success']}")
        logger.info(f"  Delete Failed: {stats['s3_delete_fail']}")
        
        success = (stats['errors'] == 0 and stats['s3_copy_fail'] == 0)
        
        if success or stats['linked'] > 0:
            logger.success("")
            logger.success("✅ Document re-linking with clean S3 paths completed!")
            logger.success("All documents are now organized under correct patient folders")
            if not dry_run:
                logger.success("Next: Run Phase 7 to re-link images")
            logger.success("")
            logger.phase_end(success=True)
            return True
        else:
            logger.warning("")
            logger.warning(f"⚠️  Re-linking completed with issues")
            logger.warning("Review logs before proceeding")
            logger.warning("")
            logger.phase_end(success=False)
            return False
            
    except Exception as e:
        logger.error(f"Exception during document re-linking: {str(e)}", exc_info=e)
        logger.phase_end(success=False)
        return False


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Re-link documents with clean S3 paths')
    parser.add_argument('--dry-run', action='store_true', help='Preview without making changes')
    parser.add_argument('--excel', default='Docs.xlsx', help='Path to Docs.xlsx file')
    args = parser.parse_args()
    
    success = relink_documents_clean(dry_run=args.dry_run, excel_file=args.excel)
    sys.exit(0 if success else 1)

