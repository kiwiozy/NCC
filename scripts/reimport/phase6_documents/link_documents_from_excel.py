#!/usr/bin/env python3
"""
Link FileMaker Documents to Patient Records (EXCEL IMPORT)

This script:
1. Reads Docs.xlsx (FileMaker export with document metadata)
2. For each document in the database:
   - Finds the corresponding patient using Doc FileMaker ID -> Patient FileMaker ID mapping
   - Updates the document's Generic Foreign Key to link to the patient
   - NO S3 OPERATIONS - documents stay where they are!
3. Fast linking - only updates database relationships

This is the "fast" version - like phase7_images/link_filemaker_images_csv.py
Documents remain in their current S3 locations (no copying/moving).

Usage:
    python3 link_documents_from_excel.py --dry-run
    python3 link_documents_from_excel.py
"""
import os
import sys
from pathlib import Path
from openpyxl import load_workbook

# Add backend to path
backend_dir = Path(__file__).parent.parent.parent.parent / 'backend'
sys.path.insert(0, str(backend_dir))

# Django setup
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ncc_api.settings')
import django
django.setup()

from documents.models import Document
from patients.models import Patient
from django.contrib.contenttypes.models import ContentType
from django.db import transaction

# Add project root for utils
project_root = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(project_root))
from scripts.reimport.utils.logger import create_logger


def link_documents_from_excel(excel_file: str = None, dry_run: bool = False) -> bool:
    """
    Link documents to patients using Excel metadata.
    
    This is the FAST version - only updates database links, no S3 operations.
    """
    logger = create_logger("PHASE 6")
    logger.phase_start("Phase 6", "Link Documents to Patients (Excel-based)")
    
    if dry_run:
        logger.warning("🔍 DRY RUN MODE - No changes will be made")
        logger.info("")
    
    # ========================================
    # Step 1: Load Excel file
    # ========================================
    if not excel_file:
        # Try common locations
        possible_paths = [
            Path.cwd() / 'Docs.xlsx',
            Path(__file__).parent.parent.parent.parent / 'Docs.xlsx',
        ]
        
        for path in possible_paths:
            if path.exists():
                excel_file = str(path)
                break
        
        if not excel_file:
            logger.error("❌ Could not find Docs.xlsx")
            logger.error("Please provide --excel-file path or place Docs.xlsx in project root")
            logger.phase_end(success=False)
            return False
    
    excel_path = Path(excel_file)
    if not excel_path.exists():
        logger.error(f"❌ Excel file not found: {excel_file}")
        logger.phase_end(success=False)
        return False
    
    logger.info(f"Reading document metadata from: {excel_file}")
    logger.info("Loading Excel workbook...")
    
    try:
        workbook = load_workbook(filename=excel_file, read_only=True, data_only=True)
        sheet = workbook.active
        
        # Get headers
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        logger.success(f"✅ Found {len(headers)} columns")
        logger.info(f"Columns: {', '.join(headers[:10])}")
        
    except Exception as e:
        logger.error(f"Failed to read Excel file: {e}")
        logger.phase_end(success=False)
        return False
    
    # ========================================
    # Step 2: Build mapping: doc_filemaker_id -> patient_filemaker_id
    # ========================================
    logger.info("")
    logger.info("Building document-to-patient mapping...")
    
    doc_to_patient_map = {}
    row_count = 0
    
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
        if not any(row):
            continue
        
        row_data = dict(zip(headers, row))
        row_count += 1
        
        doc_id = row_data.get('id')  # Document FileMaker ID
        patient_id = row_data.get('id_Contact')  # Patient FileMaker ID (corrected from 'id.key')
        
        if doc_id and patient_id:
            # Normalize to lowercase for consistent matching
            doc_to_patient_map[str(doc_id).lower()] = str(patient_id).lower()
        
        # Progress every 1000 rows
        if row_count % 1000 == 0:
            logger.info(f"💓 Processed {row_count:,} rows...")
    
    logger.success(f"✅ Built mapping for {len(doc_to_patient_map):,} documents")
    logger.info(f"Total rows in Excel: {row_count:,}")
    
    # ========================================
    # Step 3: Load all patients (for lookup)
    # ========================================
    logger.info("")
    logger.info("Loading patient lookup...")
    
    # Build patient lookup: filemaker_id -> Patient object
    patient_lookup = {}
    for patient in Patient.objects.all():
        fm_id = patient.filemaker_metadata.get('filemaker_id') if patient.filemaker_metadata else None
        if fm_id:
            # Normalize to lowercase for case-insensitive matching
            patient_lookup[str(fm_id).lower()] = patient
    
    logger.success(f"✅ Loaded {len(patient_lookup):,} patients")
    
    # ========================================
    # Step 4: Link documents to patients
    # ========================================
    logger.info("")
    logger.info("Linking documents to patients...")
    
    stats = {
        'total': 0,
        'linked': 0,
        'already_linked': 0,
        'no_mapping': 0,
        'patient_not_found': 0,
        'errors': 0,
    }
    
    # Get ContentType for Patient model
    patient_content_type = ContentType.objects.get_for_model(Patient)
    
    # Get all documents
    documents = Document.objects.all()
    total_docs = documents.count()
    logger.info(f"Found {total_docs:,} documents in database")
    
    for i, doc in enumerate(documents, 1):
        stats['total'] += 1
        
        # Progress heartbeat every 100 documents
        if i % 100 == 0:
            logger.info(f"💓 Processing document {i:,}/{total_docs:,} ({i*100/total_docs:.1f}%)")
        
        try:
            # Get document's FileMaker ID
            doc_fm_id = doc.filemaker_id
            
            if not doc_fm_id:
                stats['no_mapping'] += 1
                continue
            
            # Find patient FileMaker ID from mapping (normalize to lowercase)
            patient_fm_id = doc_to_patient_map.get(str(doc_fm_id).lower())
            
            if not patient_fm_id:
                stats['no_mapping'] += 1
                continue
            
            # Find patient
            patient = patient_lookup.get(str(patient_fm_id))
            
            if not patient:
                stats['patient_not_found'] += 1
                continue
            
            # Check if already linked to correct patient (NEW UUID check)
            if doc.content_type == patient_content_type and doc.object_id == str(patient.id):
                stats['already_linked'] += 1
                continue
            
            # Link/Re-link document to patient (handles both new links and OLD UUID updates)
            if not dry_run:
                with transaction.atomic():
                    doc.content_type = patient_content_type
                    doc.object_id = str(patient.id)
                    doc.save(update_fields=['content_type', 'object_id'])
            
            stats['linked'] += 1
            
        except Exception as e:
            logger.error(f"Error processing document {doc.id}: {e}")
            stats['errors'] += 1
    
    # ========================================
    # Summary
    # ========================================
    logger.info("")
    logger.info("=" * 70)
    logger.info("📊 Document Linking Summary")
    logger.info("=" * 70)
    logger.info(f"Total Documents: {stats['total']:,}")
    logger.info(f"✅ Linked: {stats['linked']:,}")
    logger.info(f"✓  Already Linked: {stats['already_linked']:,}")
    logger.info(f"⏭️  No Mapping: {stats['no_mapping']:,}")
    logger.info(f"⏭️  Patient Not Found: {stats['patient_not_found']:,}")
    logger.info(f"❌ Errors: {stats['errors']:,}")
    
    if dry_run:
        logger.info("")
        logger.warning("🔍 DRY RUN - No changes were made")
    
    success = stats['errors'] == 0
    
    if success:
        logger.info("")
        logger.success("✅ ✅ Document linking completed successfully!")
        logger.success(f"✅ {stats['linked']:,} documents now linked to patients")
        logger.success("Documents remain in their current S3 locations (no copying needed)")
        logger.info("")
    
    logger.phase_end(success=success)
    return success


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Link documents to patients from Excel')
    parser.add_argument('--excel-file', type=str, help='Path to Docs.xlsx')
    parser.add_argument('--dry-run', action='store_true', help='Test mode - no changes')
    
    args = parser.parse_args()
    
    success = link_documents_from_excel(
        excel_file=args.excel_file,
        dry_run=args.dry_run
    )
    
    sys.exit(0 if success else 1)

