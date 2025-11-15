#!/usr/bin/env python3
"""
Import Communications from FileMaker Excel Export

This script reads Coms.xlsx and updates patient records with:
- Phone numbers (mobile and landline)
- Email addresses
- Physical addresses

Usage:
    python import_communications.py
    python import_communications.py --dry-run
"""
import sys
import os
from pathlib import Path
from datetime import datetime
import json

# Add backend to path
backend_dir = Path(__file__).parent.parent.parent.parent / 'backend'
sys.path.insert(0, str(backend_dir))

# Django setup
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ncc_api.settings')
import django
django.setup()

from patients.models import Patient
from django.db import transaction

# Add project root for utils
project_root = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(project_root))
from scripts.reimport.utils.logger import create_logger


def import_communications(excel_file: str = "Coms.xlsx", dry_run: bool = False) -> bool:
    """
    Import communication data from Coms.xlsx and update patient records.
    
    Args:
        excel_file: Path to Coms.xlsx file
        dry_run: If True, preview changes without modifying database
    
    Returns:
        True if successful, False otherwise
    """
    logger = create_logger("COMMS")
    logger.phase_start("Phase 4.5", "Import Communication Data")
    
    if dry_run:
        logger.warning("🔍 DRY RUN MODE - No changes will be made")
        logger.info("")
    
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        logger.phase_end(success=False)
        return False
    
    # Find Excel file
    excel_path = project_root / excel_file
    if not excel_path.exists():
        logger.error(f"Excel file not found: {excel_path}")
        logger.phase_end(success=False)
        return False
    
    logger.info(f"📊 Reading communications from: {excel_path.name}")
    logger.info("")
    
    # Load Excel file
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        
        total_rows = sheet.max_row - 1
        logger.success(f"✅ Loaded {total_rows:,} communication records")
    except Exception as e:
        logger.error(f"Failed to load Excel file: {e}")
        logger.phase_end(success=False)
        return False
    
    # Group communications by patient
    logger.info("")
    logger.info("=" * 70)
    logger.info("Step 1: Grouping communications by patient...")
    logger.info("=" * 70)
    
    patient_comms = {}  # {patient_filemaker_id: {phones: [], emails: [], addresses: []}}
    
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if (i + 1) % 1000 == 0:
            logger.info(f"  💓 Processed {i + 1:,} records...")
        
        row_data = dict(zip(headers, row))
        
        patient_id = row_data.get('id.key')
        comm_type = row_data.get('type')
        
        if not patient_id or not comm_type:
            continue
        
        # Initialize patient entry if needed
        if patient_id not in patient_comms:
            patient_comms[patient_id] = {
                'phones': [],
                'emails': [],
                'addresses': []
            }
        
        # Process based on type
        if comm_type == 'Mobile':
            phone = row_data.get('ph') or row_data.get('PhoneMobileIntFormat') or row_data.get('SMS Phone')
            if phone:
                patient_comms[patient_id]['phones'].append({
                    'type': 'mobile',
                    'number': str(phone).strip(),
                    'label': row_data.get('Name') or 'Mobile'
                })
        
        elif comm_type == 'Phone':
            phone = row_data.get('ph')
            if phone:
                patient_comms[patient_id]['phones'].append({
                    'type': 'phone',
                    'number': str(phone).strip(),
                    'label': row_data.get('Name') or 'Phone'
                })
        
        elif comm_type == 'Email':
            email = row_data.get('Email default')
            if email:
                patient_comms[patient_id]['emails'].append({
                    'address': str(email).strip(),
                    'label': row_data.get('Name') or 'Email'
                })
        
        elif comm_type == 'Address':
            address_1 = row_data.get('address 1')
            if address_1:
                patient_comms[patient_id]['addresses'].append({
                    'line1': str(address_1).strip() if address_1 else '',
                    'line2': str(row_data.get('address 2')).strip() if row_data.get('address 2') else '',
                    'suburb': str(row_data.get('suburb')).strip() if row_data.get('suburb') else '',
                    'state': str(row_data.get('state')).strip() if row_data.get('state') else '',
                    'postcode': str(row_data.get('post code')).strip() if row_data.get('post code') else '',
                    'country': str(row_data.get('country')).strip() if row_data.get('country') else 'Australia',
                    'label': row_data.get('Name') or 'Address',
                    'full_address': row_data.get('Full Address one line') or ''
                })
    
    logger.success(f"✅ Grouped communications for {len(patient_comms):,} patients")
    
    # Create patient lookup map
    logger.info("")
    logger.info("=" * 70)
    logger.info("Step 2: Loading patient records...")
    logger.info("=" * 70)
    
    patients = Patient.objects.filter(
        filemaker_metadata__filemaker_id__isnull=False
    ).only('id', 'filemaker_metadata', 'contact_json', 'address_json')
    
    patient_map = {}
    for patient in patients:
        fm_id = patient.filemaker_metadata.get('filemaker_id')
        if fm_id:
            patient_map[fm_id] = patient
    
    logger.success(f"✅ Loaded {len(patient_map):,} patients")
    
    # Update patients with communication data
    logger.info("")
    logger.info("=" * 70)
    logger.info("Step 3: Updating patient records...")
    logger.info("=" * 70)
    logger.info("")
    
    stats = {
        'patients_updated': 0,
        'phones_added': 0,
        'emails_added': 0,
        'addresses_added': 0,
        'patients_not_found': 0,
        'errors': 0
    }
    
    for fm_id, comms in patient_comms.items():
        # Find patient
        patient = patient_map.get(fm_id)
        
        if not patient:
            stats['patients_not_found'] += 1
            continue
        
        # Prepare updates
        has_updates = False
        
        # Initialize JSON fields if None
        contact_json = patient.contact_json or {}
        address_json = patient.address_json or {}
        
        # Add phones
        if comms['phones']:
            contact_json['phones'] = comms['phones']
            stats['phones_added'] += len(comms['phones'])
            has_updates = True
        
        # Add emails
        if comms['emails']:
            contact_json['emails'] = comms['emails']
            # Also set primary email for convenience
            if comms['emails']:
                contact_json['email'] = comms['emails'][0]['address']
            stats['emails_added'] += len(comms['emails'])
            has_updates = True
        
        # Add addresses
        if comms['addresses']:
            address_json['addresses'] = comms['addresses']
            # Also set primary address for convenience
            if comms['addresses']:
                primary = comms['addresses'][0]
                address_json['street'] = primary['line1']
                address_json['suburb'] = primary['suburb']
                address_json['state'] = primary['state']
                address_json['postcode'] = primary['postcode']
            stats['addresses_added'] += len(comms['addresses'])
            has_updates = True
        
        # Update patient if we have changes
        if has_updates:
            if not dry_run:
                try:
                    with transaction.atomic():
                        patient.contact_json = contact_json
                        patient.address_json = address_json
                        patient.save(update_fields=['contact_json', 'address_json'])
                    stats['patients_updated'] += 1
                except Exception as e:
                    logger.error(f"Failed to update patient {patient.id}: {e}")
                    stats['errors'] += 1
            else:
                stats['patients_updated'] += 1
            
            # Progress logging
            if stats['patients_updated'] % 100 == 0:
                logger.info(f"  💓 Updated {stats['patients_updated']:,} patients...")
    
    # Summary
    logger.info("")
    logger.info("=" * 70)
    logger.info("📊 SUMMARY")
    logger.info("=" * 70)
    logger.info(f"Patients updated:      {stats['patients_updated']:,}")
    logger.info(f"Phones added:          {stats['phones_added']:,}")
    logger.info(f"Emails added:          {stats['emails_added']:,}")
    logger.info(f"Addresses added:       {stats['addresses_added']:,}")
    logger.info(f"Patients not found:    {stats['patients_not_found']:,}")
    logger.info(f"Errors:                {stats['errors']:,}")
    logger.info("")
    
    if dry_run:
        logger.warning("🔍 DRY RUN - No changes made to database")
    else:
        logger.success("✅ Communication data import complete!")
    
    logger.info("")
    logger.phase_end(success=True)
    return True


def main():
    """Main entry point."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Import communication data from Coms.xlsx'
    )
    parser.add_argument(
        '--excel-file',
        default='Coms.xlsx',
        help='Path to Coms Excel file (default: Coms.xlsx in project root)'
    )
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Preview changes without modifying database'
    )
    
    args = parser.parse_args()
    
    success = import_communications(
        excel_file=args.excel_file,
        dry_run=args.dry_run
    )
    
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()

