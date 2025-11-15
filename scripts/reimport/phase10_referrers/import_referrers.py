#!/usr/bin/env python3
"""
Import Referrers from Excel (Referrers.xlsx)

Reads referrer data from Excel file and creates Referrer and Specialty records.

Usage:
    python3 import_referrers.py --dry-run
    python3 import_referrers.py
"""
import os
import sys
from pathlib import Path
from openpyxl import load_workbook

# Add backend to path
backend_dir = Path(__file__).parent.parent.parent.parent / 'backend'
sys.path.insert(0, str(backend_dir))

# Django setup
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ncc_api.settings')
import django
django.setup()

from django.db import transaction
from referrers.models import Referrer, Specialty

# Add project root for utils
project_root = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(project_root))
from scripts.reimport.utils.logger import create_logger
from scripts.reimport.utils.progress_tracker import ProgressTracker


def import_referrers(excel_file: str = None, dry_run: bool = False) -> bool:
    logger = create_logger("PHASE 10")
    logger.phase_start("Phase 10", "Import Referrers from Excel")

    if dry_run:
        logger.warning("🔍 DRY RUN MODE - No changes will be made")
        logger.info("")

    # Find Excel file
    if not excel_file:
        possible_paths = [
            Path.cwd() / 'Referrers.xlsx',
            Path(__file__).parent.parent.parent.parent / 'Referrers.xlsx',
            Path(__file__).parent.parent.parent.parent / 'scripts' / 'Export_Filemaker' / 'Referrers.xlsx',
        ]

        for path in possible_paths:
            if path.exists():
                excel_file = str(path)
                break

        if not excel_file:
            logger.error("❌ Could not find Referrers.xlsx")
            logger.error("Please provide --excel-file path")
            logger.phase_end(success=False)
            return False

    excel_path = Path(excel_file)
    if not excel_path.exists():
        logger.error(f"❌ Excel file not found: {excel_file}")
        logger.phase_end(success=False)
        return False

    logger.info(f"Reading referrers from: {excel_file}")
    logger.info("Loading Excel workbook...")

    try:
        workbook = load_workbook(filename=excel_file, read_only=True, data_only=True)
        sheet = workbook.active

        # Get headers
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        logger.success(f"✅ Found {len(headers)} columns")

    except Exception as e:
        logger.error(f"Failed to read Excel file: {e}")
        logger.phase_end(success=False)
        return False

    # Get/create specialties cache
    specialties_cache = {}
    if not dry_run:
        for specialty in Specialty.objects.all():
            specialties_cache[specialty.name] = specialty

    # Process referrers
    logger.info("")
    logger.info("Processing referrers...")

    stats = {
        'total': 0,
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'specialties_created': 0,
        'errors': 0,
    }

    referrers_to_create = []
    referrers_to_update = []
    row_count = 0

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
        if not any(row):
            continue

        row_data = dict(zip(headers, row))
        row_count += 1
        stats['total'] += 1

        try:
            # Extract fields
            filemaker_id = row_data.get('id')
            title = row_data.get('Title') or ''
            first_name = row_data.get('First Name') or ''
            last_name = row_data.get('Last Name') or ''
            specialty_name = row_data.get('Specialty') or ''
            practice_name = row_data.get('Practice') or ''

            # Skip if no name
            if not first_name or not last_name or not filemaker_id:
                stats['skipped'] += 1
                continue

            # Get or create specialty
            specialty = None
            if specialty_name and not dry_run:
                if specialty_name not in specialties_cache:
                    specialty, created = Specialty.objects.get_or_create(name=specialty_name)
                    specialties_cache[specialty_name] = specialty
                    if created:
                        stats['specialties_created'] += 1
                else:
                    specialty = specialties_cache[specialty_name]

            # Build contact JSON
            contact_json = {}
            if row_data.get('Phone'):
                contact_json['phone'] = str(row_data.get('Phone'))
            if row_data.get('Mobile'):
                contact_json['mobile'] = str(row_data.get('Mobile'))
            if row_data.get('Email'):
                contact_json['email'] = str(row_data.get('Email'))

            # Build address JSON
            address_json = {}
            if row_data.get('Street'):
                address_json['street'] = str(row_data.get('Street'))
            if row_data.get('Street2'):
                address_json['street2'] = str(row_data.get('Street2'))
            if row_data.get('Suburb'):
                address_json['suburb'] = str(row_data.get('Suburb'))
            if row_data.get('State'):
                address_json['state'] = str(row_data.get('State'))
            if row_data.get('Postcode'):
                address_json['postcode'] = str(row_data.get('Postcode'))

            # Check if referrer exists
            existing_referrer = None
            if filemaker_id:
                try:
                    existing_referrer = Referrer.objects.get(filemaker_id=filemaker_id)
                except Referrer.DoesNotExist:
                    pass

            if existing_referrer:
                # Update existing
                existing_referrer.title = title
                existing_referrer.first_name = first_name
                existing_referrer.last_name = last_name
                existing_referrer.specialty = specialty
                existing_referrer.practice_name = practice_name
                existing_referrer.contact_json = contact_json or {}
                existing_referrer.address_json = address_json or {}
                referrers_to_update.append(existing_referrer)
                stats['updated'] += 1
            else:
                # Create new
                referrer = Referrer(
                    title=title,
                    first_name=first_name,
                    last_name=last_name,
                    specialty=specialty,
                    practice_name=practice_name,
                    contact_json=contact_json or {},
                    address_json=address_json or {},
                    filemaker_id=filemaker_id,
                )
                referrers_to_create.append(referrer)
                stats['created'] += 1

        except Exception as e:
            logger.error(f"Error processing row {row_count}: {e}")
            stats['errors'] += 1

        # Progress every 10 rows
        if row_count % 10 == 0:
            logger.info(f"💓 Processed {row_count:,} rows...")

    logger.success(f"✅ Processed {row_count:,} rows")

    # Save to database
    if not dry_run:
        logger.info("")
        logger.info("Saving to database...")

        with transaction.atomic():
            if referrers_to_create:
                Referrer.objects.bulk_create(referrers_to_create, batch_size=100)
                logger.success(f"✅ Created {len(referrers_to_create):,} referrers")

            if referrers_to_update:
                Referrer.objects.bulk_update(
                    referrers_to_update,
                    ['title', 'first_name', 'last_name', 'specialty', 'practice_name', 'contact_json', 'address_json'],
                    batch_size=100
                )
                logger.success(f"✅ Updated {len(referrers_to_update):,} referrers")

    # Summary
    logger.info("")
    logger.info("=" * 70)
    logger.info("📊 Referrers Import Summary")
    logger.info("=" * 70)
    logger.info(f"Total Rows: {stats['total']:,}")
    logger.success(f"✅ Created: {stats['created']:,}")
    logger.success(f"✅ Updated: {stats['updated']:,}")
    logger.success(f"✅ Specialties Created: {stats['specialties_created']:,}")
    logger.info(f"⏭️  Skipped: {stats['skipped']:,}")
    logger.error(f"❌ Errors: {stats['errors']:,}")
    logger.info("")

    if stats['errors'] == 0:
        logger.success("✅ ✅ ✅ Referrers import completed successfully!")
        success = True
    else:
        logger.error("❌ Referrers import completed with errors!")
        success = False

    logger.phase_end(success=success)
    return success


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Import referrers from Excel')
    parser.add_argument('--excel-file', help='Path to Referrers.xlsx')
    parser.add_argument('--dry-run', action='store_true', help='Dry run mode')
    args = parser.parse_args()

    success = import_referrers(excel_file=args.excel_file, dry_run=args.dry_run)
    sys.exit(0 if success else 1)

