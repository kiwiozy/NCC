#!/usr/bin/env python3
"""
Import Communications for Referrers and Companies from Coms.xlsx

This script:
1. Reads Coms.xlsx (communications table)
2. Matches communication records to Referrers and Companies by FileMaker ID
3. Updates their contact_json fields with phones, emails, addresses

Usage:
    cd /Users/craig/Documents/nexus-core-clinic/scripts/reimport
    python3 phase10_referrers/import_referrer_company_communications.py --dry-run
    python3 phase10_referrers/import_referrer_company_communications.py
"""
import os
import sys
from pathlib import Path
from openpyxl import load_workbook
import uuid

# Add backend to path
backend_dir = Path(__file__).parent.parent.parent.parent / 'backend'
sys.path.insert(0, str(backend_dir))

# Django setup
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ncc_api.settings')
import django
django.setup()

from django.db import transaction
from referrers.models import Referrer
from companies.models import Company

# Add project root for utils
project_root = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(project_root))
from scripts.reimport.utils.logger import create_logger
from scripts.reimport.utils.progress_tracker import ProgressTracker


def import_referrer_company_communications(excel_file: str = None, dry_run: bool = False) -> bool:
    """
    Import communications for referrers and companies from Coms.xlsx
    
    Args:
        excel_file: Path to Coms.xlsx file
        dry_run: If True, only preview data without saving
    
    Returns:
        True if successful, False otherwise
    """
    logger = create_logger("PHASE 10.5")
    logger.phase_start("Phase 10.5", "Import Referrer & Company Communications")

    if dry_run:
        logger.warning("🔍 DRY RUN MODE - No changes will be made")
        logger.info("")

    # ========================================
    # Step 1: Find Excel file
    # ========================================
    if not excel_file:
        possible_paths = [
            Path.cwd() / 'scripts/Export_Filemaker/Coms.xlsx',
            Path(__file__).parent.parent.parent.parent / 'scripts/Export_Filemaker/Coms.xlsx',
        ]

        for path in possible_paths:
            if path.exists():
                excel_file = str(path)
                break

        if not excel_file:
            logger.error("❌ Could not find Coms.xlsx")
            logger.error("Please provide --excel-file path or place Coms.xlsx in scripts/Export_Filemaker/")
            logger.phase_end(success=False)
            return False

    excel_path = Path(excel_file)
    if not excel_path.exists():
        logger.error(f"❌ Excel file not found: {excel_file}")
        logger.phase_end(success=False)
        return False

    logger.info(f"Reading communications from: {excel_file}")
    logger.info("Loading Excel workbook...")

    try:
        workbook = load_workbook(filename=excel_file, read_only=True, data_only=True)
        sheet = workbook.active

        # Get headers
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        logger.success(f"✅ Found {len(headers)} columns")

    except Exception as e:
        logger.error(f"Failed to read Excel file: {e}")
        logger.phase_end(success=False)
        return False

    # ========================================
    # Step 2: Build lookups for Referrers and Companies
    # ========================================
    logger.info("")
    logger.info("Building Referrer and Company lookups...")

    referrer_lookup = {}
    for referrer in Referrer.objects.filter(filemaker_id__isnull=False):
        referrer_lookup[str(referrer.filemaker_id).lower()] = referrer

    company_lookup = {}
    for company in Company.objects.filter(filemaker_id__isnull=False):
        company_lookup[str(company.filemaker_id).lower()] = company

    logger.success(f"✅ Loaded {len(referrer_lookup):,} referrers and {len(company_lookup):,} companies for lookup")

    # ========================================
    # Step 3: Process communications
    # ========================================
    logger.info("")
    logger.info("Processing communications...")

    # Group communications by contact ID
    referrer_coms = {}  # {referrer_id: [com_records]}
    company_coms = {}   # {company_id: [com_records]}
    
    row_count = 0
    stats = {
        'total_rows': 0,
        'referrer_coms': 0,
        'company_coms': 0,
        'skipped': 0,
    }

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
        if not any(row):
            continue

        row_data = dict(zip(headers, row))
        stats['total_rows'] += 1
        row_count += 1

        # Get contact ID
        contact_id_str = row_data.get('id_Contact')
        if not contact_id_str:
            stats['skipped'] += 1
            continue

        try:
            contact_id = str(contact_id_str).lower()
        except (ValueError, AttributeError):
            stats['skipped'] += 1
            continue

        # Build communication record
        com_record = {
            'type': row_data.get('Type') or '',
            'value': row_data.get('ph') or '',  # Phone/Email value
            'label': row_data.get('label') or '',
            'is_default': str(row_data.get('default')).lower() == '1' if row_data.get('default') else False,
        }

        # Check if this is a referrer
        if contact_id in referrer_lookup:
            if contact_id not in referrer_coms:
                referrer_coms[contact_id] = []
            referrer_coms[contact_id].append(com_record)
            stats['referrer_coms'] += 1

        # Check if this is a company
        elif contact_id in company_lookup:
            if contact_id not in company_coms:
                company_coms[contact_id] = []
            company_coms[contact_id].append(com_record)
            stats['company_coms'] += 1
        else:
            stats['skipped'] += 1

        # Progress every 100 rows
        if row_count % 100 == 0:
            logger.info(f"💓 Processed {row_count:,} communication records...")

    logger.success(f"✅ Processed {row_count:,} communication records")
    logger.info(f"   Referrer communications: {stats['referrer_coms']:,}")
    logger.info(f"   Company communications: {stats['company_coms']:,}")
    logger.info(f"   Skipped: {stats['skipped']:,}")

    # ========================================
    # Step 4: Update Referrers
    # ========================================
    if not dry_run:
        logger.info("")
        logger.info("Updating Referrers with communications...")

        referrers_updated = 0
        with transaction.atomic():
            for contact_id, coms in referrer_coms.items():
                referrer = referrer_lookup[contact_id]
                
                # Build contact_json
                contact_json = referrer.contact_json or {}
                phones = []
                emails = []

                for com in coms:
                    com_type = com['type'].lower()
                    value = com['value'].strip() if com['value'] else ''
                    
                    if not value:
                        continue

                    if 'phone' in com_type or 'mobile' in com_type:
                        phones.append({
                            'type': 'mobile' if 'mobile' in com_type else 'phone',
                            'number': value,
                            'label': com['label'] or ('Mobile' if 'mobile' in com_type else 'Phone'),
                            'is_default': com['is_default']
                        })
                    elif 'email' in com_type:
                        emails.append({
                            'address': value,
                            'label': com['label'] or 'Email',
                            'is_default': com['is_default']
                        })

                # Update contact_json
                if phones:
                    contact_json['phones'] = phones
                if emails:
                    contact_json['emails'] = emails

                referrer.contact_json = contact_json
                referrer.save()
                referrers_updated += 1

        logger.success(f"✅ Updated {referrers_updated:,} referrers with communications")

        # ========================================
        # Step 5: Update Companies
        # ========================================
        logger.info("")
        logger.info("Updating Companies with communications...")

        companies_updated = 0
        with transaction.atomic():
            for contact_id, coms in company_coms.items():
                company = company_lookup[contact_id]
                
                # Build contact_json
                contact_json = company.contact_json or {}
                phones = []
                emails = []

                for com in coms:
                    com_type = com['type'].lower()
                    value = com['value'].strip() if com['value'] else ''
                    
                    if not value:
                        continue

                    if 'phone' in com_type or 'mobile' in com_type or 'fax' in com_type:
                        phones.append({
                            'type': 'fax' if 'fax' in com_type else ('mobile' if 'mobile' in com_type else 'phone'),
                            'number': value,
                            'label': com['label'] or ('Fax' if 'fax' in com_type else ('Mobile' if 'mobile' in com_type else 'Phone')),
                            'is_default': com['is_default']
                        })
                    elif 'email' in com_type:
                        emails.append({
                            'address': value,
                            'label': com['label'] or 'Email',
                            'is_default': com['is_default']
                        })

                # Update contact_json
                if phones:
                    contact_json['phones'] = phones
                if emails:
                    contact_json['emails'] = emails

                company.contact_json = contact_json
                company.save()
                companies_updated += 1

        logger.success(f"✅ Updated {companies_updated:,} companies with communications")

    else:
        # Dry run - show preview
        logger.info("")
        logger.info("=" * 70)
        logger.info("🔍 DRY RUN - Communication Preview")
        logger.info("=" * 70)
        logger.info(f"Would update {len(referrer_coms)} referrers")
        logger.info(f"Would update {len(company_coms)} companies")
        logger.info("")

    # ========================================
    # Summary
    # ========================================
    logger.info("")
    logger.info("=" * 70)
    logger.info("📊 Communications Import Summary")
    logger.info("=" * 70)
    logger.info(f"Total rows processed: {stats['total_rows']:,}")
    logger.info(f"Referrer communications: {stats['referrer_coms']:,}")
    logger.info(f"Company communications: {stats['company_coms']:,}")
    if not dry_run:
        logger.success(f"✅ Referrers updated: {referrers_updated:,}")
        logger.success(f"✅ Companies updated: {companies_updated:,}")
    logger.info(f"⏭️  Skipped: {stats['skipped']:,}")
    logger.info("")

    if dry_run or (not dry_run and referrers_updated >= 0 and companies_updated >= 0):
        logger.success("✅ ✅ ✅ Communications import completed successfully!")
        success = True
    else:
        logger.error("❌ Communications import completed with errors!")
        success = False

    logger.phase_end(success=success)
    return success


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Import communications for referrers and companies from Coms.xlsx')
    parser.add_argument('--excel-file', type=str, help='Path to the Coms.xlsx file')
    parser.add_argument('--dry-run', action='store_true', help='Preview only, do not save')
    args = parser.parse_args()

    success = import_referrer_company_communications(excel_file=args.excel_file, dry_run=args.dry_run)
    sys.exit(0 if success else 1)

