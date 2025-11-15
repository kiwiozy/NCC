#!/usr/bin/env python3
"""
Import Patient-Referrer Relationships from Excel (PatientToReferrer.xlsx)

Reads patient-referrer relationships from Excel file and creates PatientReferrer records.

Usage:
    python3 import_patient_referrers.py --dry-run
    python3 import_patient_referrers.py
"""
import os
import sys
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

# Add backend to path
backend_dir = Path(__file__).parent.parent.parent.parent / 'backend'
sys.path.insert(0, str(backend_dir))

# Django setup
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ncc_api.settings')
import django
django.setup()

from django.db import transaction
from referrers.models import Referrer, PatientReferrer
from patients.models import Patient

# Add project root for utils
project_root = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(project_root))
from scripts.reimport.utils.logger import create_logger
from scripts.reimport.utils.progress_tracker import ProgressTracker


def import_patient_referrers(excel_file: str = None, dry_run: bool = False) -> bool:
    logger = create_logger("PHASE 10.5")
    logger.phase_start("Phase 10.5", "Import Patient-Referrer Relationships")

    if dry_run:
        logger.warning("🔍 DRY RUN MODE - No changes will be made")
        logger.info("")

    # Find Excel file
    if not excel_file:
        possible_paths = [
            Path.cwd() / 'PatientToReferrer.xlsx',
            Path(__file__).parent.parent.parent.parent / 'PatientToReferrer.xlsx',
            Path(__file__).parent.parent.parent.parent / 'scripts' / 'Export_Filemaker' / 'PatientToReferrer.xlsx',
        ]

        for path in possible_paths:
            if path.exists():
                excel_file = str(path)
                break

        if not excel_file:
            logger.error("❌ Could not find PatientToReferrer.xlsx")
            logger.error("Please provide --excel-file path")
            logger.phase_end(success=False)
            return False

    excel_path = Path(excel_file)
    if not excel_path.exists():
        logger.error(f"❌ Excel file not found: {excel_file}")
        logger.phase_end(success=False)
        return False

    logger.info(f"Reading patient-referrer relationships from: {excel_file}")
    logger.info("Loading Excel workbook...")

    try:
        workbook = load_workbook(filename=excel_file, read_only=True, data_only=True)
        sheet = workbook.active

        # Get headers
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        logger.success(f"✅ Found {len(headers)} columns")

    except Exception as e:
        logger.error(f"Failed to read Excel file: {e}")
        logger.phase_end(success=False)
        return False

    # Build lookups
    logger.info("")
    logger.info("Building patient and referrer lookups...")
    
    patient_lookup = {}
    for patient in Patient.objects.all():
        fm_id = patient.filemaker_metadata.get('filemaker_id') if patient.filemaker_metadata else None
        if fm_id:
            patient_lookup[str(fm_id).lower()] = patient

    referrer_lookup = {}
    for referrer in Referrer.objects.all():
        if referrer.filemaker_id:
            referrer_lookup[str(referrer.filemaker_id).lower()] = referrer

    logger.success(f"✅ Loaded {len(patient_lookup):,} patients")
    logger.success(f"✅ Loaded {len(referrer_lookup):,} referrers")

    # Process relationships
    logger.info("")
    logger.info("Processing patient-referrer relationships...")

    stats = {
        'total': 0,
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'patient_not_found': 0,
        'referrer_not_found': 0,
        'errors': 0,
    }

    relationships_to_create = []
    relationships_to_update = []
    row_count = 0

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
        if not any(row):
            continue

        row_data = dict(zip(headers, row))
        row_count += 1
        stats['total'] += 1

        try:
            # Extract fields
            filemaker_id = row_data.get('id')
            patient_fm_id = row_data.get('id_Contact')
            referrer_fm_id = row_data.get('id_Perscriber')  # Fixed: Typo in FileMaker - it's 'Perscriber' not 'Referrer'
            referral_date = row_data.get('date')  # Fixed: lowercase
            referral_reason = row_data.get('Reason') or ''
            status = row_data.get('Status') or 'ACTIVE'

            # Map status
            status_mapping = {
                'Active': 'ACTIVE',
                'Inactive': 'INACTIVE',
                'Pending': 'PENDING',
            }
            status = status_mapping.get(status, 'ACTIVE')

            # Parse date
            parsed_date = None
            if referral_date:
                if isinstance(referral_date, datetime):
                    parsed_date = referral_date.date()
                elif isinstance(referral_date, str):
                    try:
                        parsed_date = datetime.strptime(referral_date, '%Y-%m-%d').date()
                    except:
                        pass

            # Find patient
            patient = patient_lookup.get(str(patient_fm_id).lower()) if patient_fm_id else None
            if not patient:
                stats['patient_not_found'] += 1
                stats['skipped'] += 1
                continue

            # Find referrer
            referrer = referrer_lookup.get(str(referrer_fm_id).lower()) if referrer_fm_id else None
            if not referrer:
                stats['referrer_not_found'] += 1
                stats['skipped'] += 1
                continue

            # Check if relationship exists
            existing_rel = None
            if filemaker_id:
                try:
                    existing_rel = PatientReferrer.objects.get(filemaker_id=filemaker_id)
                except PatientReferrer.DoesNotExist:
                    pass

            if not existing_rel:
                # Check by patient+referrer
                try:
                    existing_rel = PatientReferrer.objects.get(patient=patient, referrer=referrer)
                except PatientReferrer.DoesNotExist:
                    pass

            if existing_rel:
                # Update existing
                existing_rel.referral_date = parsed_date
                existing_rel.referral_reason = referral_reason
                existing_rel.status = status
                relationships_to_update.append(existing_rel)
                stats['updated'] += 1
            else:
                # Create new
                rel = PatientReferrer(
                    patient=patient,
                    referrer=referrer,
                    referral_date=parsed_date,
                    referral_reason=referral_reason,
                    status=status,
                    filemaker_id=filemaker_id,
                )
                relationships_to_create.append(rel)
                stats['created'] += 1

        except Exception as e:
            logger.error(f"Error processing row {row_count}: {e}")
            stats['errors'] += 1

        # Progress every 25 rows
        if row_count % 25 == 0:
            logger.info(f"💓 Processed {row_count:,} rows...")

    logger.success(f"✅ Processed {row_count:,} rows")

    # Save to database
    if not dry_run:
        logger.info("")
        logger.info("Saving to database...")

        with transaction.atomic():
            if relationships_to_create:
                PatientReferrer.objects.bulk_create(relationships_to_create, batch_size=100, ignore_conflicts=True)  # Fixed: Ignore duplicates
                logger.success(f"✅ Created {len(relationships_to_create):,} relationships (duplicates ignored)")

            if relationships_to_update:
                PatientReferrer.objects.bulk_update(
                    relationships_to_update,
                    ['referral_date', 'referral_reason', 'status'],
                    batch_size=100
                )
                logger.success(f"✅ Updated {len(relationships_to_update):,} relationships")

        # Set the most recent referrer as primary for each patient
        logger.info("")
        logger.info("Setting primary referrers for each patient...")
        
        patients_with_referrers = Patient.objects.filter(patient_referrers__isnull=False).distinct()
        primary_count = 0
        
        for patient in patients_with_referrers:
            # Get all active referrers for this patient, ordered by most recent date
            patient_referrers = PatientReferrer.objects.filter(
                patient=patient,
                status='ACTIVE'
            ).order_by('-referral_date', '-created_at')
            
            if patient_referrers.exists():
                # Set the first (most recent) as primary
                most_recent = patient_referrers.first()
                most_recent.is_primary = True
                most_recent.save(update_fields=['is_primary'])
                
                # Set all others as non-primary
                patient_referrers.exclude(id=most_recent.id).update(is_primary=False)
                primary_count += 1
        
        logger.success(f"✅ Set primary referrers for {primary_count:,} patients")

    # Summary
    logger.info("")
    logger.info("=" * 70)
    logger.info("📊 Patient-Referrer Relationships Import Summary")
    logger.info("=" * 70)
    logger.info(f"Total Rows: {stats['total']:,}")
    logger.success(f"✅ Created: {stats['created']:,}")
    logger.success(f"✅ Updated: {stats['updated']:,}")
    logger.info(f"⏭️  Skipped: {stats['skipped']:,}")
    logger.info(f"   - Patient not found: {stats['patient_not_found']:,}")
    logger.info(f"   - Referrer not found: {stats['referrer_not_found']:,}")
    logger.error(f"❌ Errors: {stats['errors']:,}")
    logger.info("")

    if stats['errors'] == 0:
        logger.success("✅ ✅ ✅ Patient-Referrer relationships import completed successfully!")
        success = True
    else:
        logger.error("❌ Patient-Referrer relationships import completed with errors!")
        success = False

    logger.phase_end(success=success)
    return success


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Import patient-referrer relationships from Excel')
    parser.add_argument('--excel-file', help='Path to PatientToReferrer.xlsx')
    parser.add_argument('--dry-run', action='store_true', help='Dry run mode')
    args = parser.parse_args()

    success = import_patient_referrers(excel_file=args.excel_file, dry_run=args.dry_run)
    sys.exit(0 if success else 1)

