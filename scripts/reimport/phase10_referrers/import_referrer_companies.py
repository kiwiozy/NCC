#!/usr/bin/env python3
"""
Import Referrer-Company Relationships from Excel (ReferrerToCompanies.xlsx)

Reads referrer-company relationships from Excel file and creates ReferrerCompany records.

Usage:
    python3 import_referrer_companies.py --dry-run
    python3 import_referrer_companies.py
"""
import os
import sys
from pathlib import Path
from openpyxl import load_workbook

# Add backend to path
backend_dir = Path(__file__).parent.parent.parent.parent / 'backend'
sys.path.insert(0, str(backend_dir))

# Django setup
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ncc_api.settings')
import django
django.setup()

from django.db import transaction
from referrers.models import Referrer, ReferrerCompany
from companies.models import Company

# Add project root for utils
project_root = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(project_root))
from scripts.reimport.utils.logger import create_logger
from scripts.reimport.utils.progress_tracker import ProgressTracker


def import_referrer_companies(excel_file: str = None, dry_run: bool = False) -> bool:
    logger = create_logger("PHASE 10.6")
    logger.phase_start("Phase 10.6", "Import Referrer-Company Relationships")

    if dry_run:
        logger.warning("🔍 DRY RUN MODE - No changes will be made")
        logger.info("")

    # Find Excel file
    if not excel_file:
        possible_paths = [
            Path.cwd() / 'ReferrerToCompanies.xlsx',
            Path(__file__).parent.parent.parent.parent / 'ReferrerToCompanies.xlsx',
            Path(__file__).parent.parent.parent.parent / 'scripts' / 'Export_Filemaker' / 'ReferrerToCompanies.xlsx',
        ]

        for path in possible_paths:
            if path.exists():
                excel_file = str(path)
                break

        if not excel_file:
            logger.error("❌ Could not find ReferrerToCompanies.xlsx")
            logger.error("Please provide --excel-file path")
            logger.phase_end(success=False)
            return False

    excel_path = Path(excel_file)
    if not excel_path.exists():
        logger.error(f"❌ Excel file not found: {excel_file}")
        logger.phase_end(success=False)
        return False

    logger.info(f"Reading referrer-company relationships from: {excel_file}")
    logger.info("Loading Excel workbook...")

    try:
        workbook = load_workbook(filename=excel_file, read_only=True, data_only=True)
        sheet = workbook.active

        # Get headers
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        logger.success(f"✅ Found {len(headers)} columns")

    except Exception as e:
        logger.error(f"Failed to read Excel file: {e}")
        logger.phase_end(success=False)
        return False

    # Build lookups
    logger.info("")
    logger.info("Building referrer and company lookups...")
    
    referrer_lookup = {}
    for referrer in Referrer.objects.all():
        if referrer.filemaker_id:
            referrer_lookup[str(referrer.filemaker_id).lower()] = referrer

    company_lookup = {}
    for company in Company.objects.all():
        if company.filemaker_id:
            company_lookup[str(company.filemaker_id).lower()] = company

    logger.success(f"✅ Loaded {len(referrer_lookup):,} referrers")
    logger.success(f"✅ Loaded {len(company_lookup):,} companies")

    # Process relationships
    logger.info("")
    logger.info("Processing referrer-company relationships...")

    stats = {
        'total': 0,
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'referrer_not_found': 0,
        'company_not_found': 0,
        'errors': 0,
    }

    relationships_to_create = []
    relationships_to_update = []
    row_count = 0

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
        if not any(row):
            continue

        row_data = dict(zip(headers, row))
        row_count += 1
        stats['total'] += 1

        try:
            # Extract fields
            filemaker_id = row_data.get('id')
            referrer_fm_id = row_data.get('id_Referrer')
            company_fm_id = row_data.get('id_Company')
            position = row_data.get('Position') or ''
            is_primary = row_data.get('Primary') or False

            # Handle boolean conversion
            if isinstance(is_primary, str):
                is_primary = is_primary.lower() in ('yes', 'true', '1')
            else:
                is_primary = bool(is_primary)

            # Find referrer
            referrer = referrer_lookup.get(str(referrer_fm_id).lower()) if referrer_fm_id else None
            if not referrer:
                stats['referrer_not_found'] += 1
                stats['skipped'] += 1
                continue

            # Find company
            company = company_lookup.get(str(company_fm_id).lower()) if company_fm_id else None
            if not company:
                stats['company_not_found'] += 1
                stats['skipped'] += 1
                continue

            # Check if relationship exists
            existing_rel = None
            if filemaker_id:
                try:
                    existing_rel = ReferrerCompany.objects.get(filemaker_id=filemaker_id)
                except ReferrerCompany.DoesNotExist:
                    pass

            if not existing_rel:
                # Check by referrer+company
                try:
                    existing_rel = ReferrerCompany.objects.get(referrer=referrer, company=company)
                except ReferrerCompany.DoesNotExist:
                    pass

            if existing_rel:
                # Update existing
                existing_rel.position = position
                existing_rel.is_primary = is_primary
                relationships_to_update.append(existing_rel)
                stats['updated'] += 1
            else:
                # Create new
                rel = ReferrerCompany(
                    referrer=referrer,
                    company=company,
                    position=position,
                    is_primary=is_primary,
                    filemaker_id=filemaker_id,
                )
                relationships_to_create.append(rel)
                stats['created'] += 1

        except Exception as e:
            logger.error(f"Error processing row {row_count}: {e}")
            stats['errors'] += 1

        # Progress every 10 rows
        if row_count % 10 == 0:
            logger.info(f"💓 Processed {row_count:,} rows...")

    logger.success(f"✅ Processed {row_count:,} rows")

    # Save to database
    if not dry_run:
        logger.info("")
        logger.info("Saving to database...")

        with transaction.atomic():
            if relationships_to_create:
                ReferrerCompany.objects.bulk_create(relationships_to_create, batch_size=100)
                logger.success(f"✅ Created {len(relationships_to_create):,} relationships")

            if relationships_to_update:
                ReferrerCompany.objects.bulk_update(
                    relationships_to_update,
                    ['position', 'is_primary'],
                    batch_size=100
                )
                logger.success(f"✅ Updated {len(relationships_to_update):,} relationships")

    # Summary
    logger.info("")
    logger.info("=" * 70)
    logger.info("📊 Referrer-Company Relationships Import Summary")
    logger.info("=" * 70)
    logger.info(f"Total Rows: {stats['total']:,}")
    logger.success(f"✅ Created: {stats['created']:,}")
    logger.success(f"✅ Updated: {stats['updated']:,}")
    logger.info(f"⏭️  Skipped: {stats['skipped']:,}")
    logger.info(f"   - Referrer not found: {stats['referrer_not_found']:,}")
    logger.info(f"   - Company not found: {stats['company_not_found']:,}")
    logger.error(f"❌ Errors: {stats['errors']:,}")
    logger.info("")

    if stats['errors'] == 0:
        logger.success("✅ ✅ ✅ Referrer-Company relationships import completed successfully!")
        success = True
    else:
        logger.error("❌ Referrer-Company relationships import completed with errors!")
        success = False

    logger.phase_end(success=success)
    return success


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Import referrer-company relationships from Excel')
    parser.add_argument('--excel-file', help='Path to ReferrerToCompanies.xlsx')
    parser.add_argument('--dry-run', action='store_true', help='Dry run mode')
    args = parser.parse_args()

    success = import_referrer_companies(excel_file=args.excel_file, dry_run=args.dry_run)
    sys.exit(0 if success else 1)

