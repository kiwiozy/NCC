#!/usr/bin/env python3
"""
Extract and Import Coordinators from Contacts.xlsx

This script runs BEFORE patient import to:
1. Scan all patient records in Contacts.xlsx
2. Extract unique NDIS coordinator names, phones, emails, notes
3. Merge contact details for coordinators mentioned by multiple patients
4. Create Referrer records with specialty = "Support Coordinator"

This allows patients to be linked to coordinators during patient import.

Usage:
    cd /Users/craig/Documents/nexus-core-clinic/scripts/reimport
    python3 phase2_coordinators/extract_coordinators_from_contacts.py --dry-run
    python3 phase2_coordinators/extract_coordinators_from_contacts.py
"""
import os
import sys
from pathlib import Path
from openpyxl import load_workbook
import uuid

# Add backend to path
backend_dir = Path(__file__).parent.parent.parent.parent / 'backend'
sys.path.insert(0, str(backend_dir))

# Django setup
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ncc_api.settings')
import django
django.setup()

from django.db import transaction
from referrers.models import Referrer, Specialty

# Add project root for utils
project_root = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(project_root))
from scripts.reimport.utils.logger import create_logger
from scripts.reimport.utils.progress_tracker import ProgressTracker


def split_name(full_name: str) -> tuple:
    """Split full name into first and last name"""
    if not full_name:
        return "", ""
    
    parts = full_name.strip().split()
    if len(parts) == 0:
        return "", ""
    elif len(parts) == 1:
        return parts[0], ""
    else:
        first_name = parts[0]
        last_name = " ".join(parts[1:])
        return first_name, last_name


def extract_coordinators_from_contacts(excel_file: str = None, dry_run: bool = False) -> bool:
    """
    Extract unique coordinators from Contacts.xlsx and create Referrer records.
    
    Args:
        excel_file: Path to Contacts.xlsx file
        dry_run: If True, only preview data without saving
    
    Returns:
        True if successful, False otherwise
    """
    logger = create_logger("PHASE 2.5")
    logger.phase_start("Phase 2.5", "Extract Coordinators from Contacts.xlsx")

    if dry_run:
        logger.warning("🔍 DRY RUN MODE - No changes will be made")
        logger.info("")

    # ========================================
    # Step 1: Find Excel file
    # ========================================
    if not excel_file:
        possible_paths = [
            Path.cwd() / 'scripts/Export_Filemaker/Contacts.xlsx',
            Path(__file__).parent.parent.parent.parent / 'scripts/Export_Filemaker/Contacts.xlsx',
        ]

        for path in possible_paths:
            if path.exists():
                excel_file = str(path)
                break

        if not excel_file:
            logger.error("❌ Could not find Contacts.xlsx")
            logger.error("Please provide --excel-file path or place Contacts.xlsx in scripts/Export_Filemaker/")
            logger.phase_end(success=False)
            return False

    excel_path = Path(excel_file)
    if not excel_path.exists():
        logger.error(f"❌ Excel file not found: {excel_file}")
        logger.phase_end(success=False)
        return False

    logger.info(f"Reading patient data from: {excel_file}")
    logger.info("Loading Excel workbook...")

    try:
        workbook = load_workbook(filename=excel_file, read_only=True, data_only=True)
        sheet = workbook.active

        # Get headers
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        logger.success(f"✅ Found {len(headers)} columns")

    except Exception as e:
        logger.error(f"Failed to read Excel file: {e}")
        logger.phase_end(success=False)
        return False

    # ========================================
    # Step 2: Extract unique coordinators
    # ========================================
    logger.info("")
    logger.info("Extracting unique coordinators...")

    unique_coordinators = {}
    row_count = 0

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
        if not any(row):
            continue

        row_data = dict(zip(headers, row))
        row_count += 1

        # Extract coordinator fields
        coordinator_name = row_data.get('NDIS Coordinator Name')
        coordinator_phone = row_data.get('NDIS Coordinator Phone')
        coordinator_email = row_data.get('NDIS Coordinator Email')
        ndis_notes = row_data.get('NDIS notes')

        # Skip if no coordinator name
        if not coordinator_name or not str(coordinator_name).strip():
            continue

        coordinator_name = str(coordinator_name).strip()

        # Initialize coordinator entry if new
        if coordinator_name not in unique_coordinators:
            unique_coordinators[coordinator_name] = {
                'name': coordinator_name,
                'phones': [],
                'emails': [],
                'notes': []
            }

        # Add phone if present and not duplicate
        if coordinator_phone and str(coordinator_phone).strip():
            phone = str(coordinator_phone).strip()
            if phone not in unique_coordinators[coordinator_name]['phones']:
                unique_coordinators[coordinator_name]['phones'].append(phone)

        # Add email if present and not duplicate
        if coordinator_email and str(coordinator_email).strip():
            email = str(coordinator_email).strip()
            if email not in unique_coordinators[coordinator_name]['emails']:
                unique_coordinators[coordinator_name]['emails'].append(email)

        # Add notes if present
        if ndis_notes and str(ndis_notes).strip():
            note = str(ndis_notes).strip()
            if note and note not in unique_coordinators[coordinator_name]['notes']:
                unique_coordinators[coordinator_name]['notes'].append(note)

        # Progress every 100 rows
        if row_count % 100 == 0:
            logger.info(f"💓 Processed {row_count:,} patient records...")

    logger.success(f"✅ Processed {row_count:,} patient records")
    logger.success(f"✅ Found {len(unique_coordinators)} unique coordinators")

    # ========================================
    # Step 3: Create Referrer records
    # ========================================
    if not dry_run:
        logger.info("")
        logger.info("Creating Referrer records for coordinators...")

        # Get or create "Support Coordinator" specialty
        try:
            specialty, created = Specialty.objects.get_or_create(name="Support Coordinator")
            if created:
                logger.success("✅ Created 'Support Coordinator' specialty")
        except Exception as e:
            logger.error(f"Failed to create specialty: {e}")
            logger.phase_end(success=False)
            return False

        stats = {
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'errors': 0,
        }

        with transaction.atomic():
            for coordinator_name, coordinator_data in unique_coordinators.items():
                try:
                    # Split name
                    first_name, last_name = split_name(coordinator_name)

                    if not first_name or not last_name:
                        logger.warning(f"Skipping '{coordinator_name}' - could not split name")
                        stats['skipped'] += 1
                        continue

                    # Build contact_json
                    contact_json = {}
                    if coordinator_data['phones']:
                        if len(coordinator_data['phones']) == 1:
                            contact_json['phone'] = coordinator_data['phones'][0]
                        else:
                            contact_json['phones'] = coordinator_data['phones']
                    
                    if coordinator_data['emails']:
                        if len(coordinator_data['emails']) == 1:
                            contact_json['email'] = coordinator_data['emails'][0]
                        else:
                            contact_json['emails'] = coordinator_data['emails']

                    # Combine notes with separators
                    combined_notes = "\n---\n".join(coordinator_data['notes']) if coordinator_data['notes'] else None

                    # Create or update Referrer
                    referrer, created = Referrer.objects.update_or_create(
                        first_name__iexact=first_name,
                        last_name__iexact=last_name,
                        specialty=specialty,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'contact_json': contact_json,
                            'notes': combined_notes,
                        }
                    )

                    if created:
                        stats['created'] += 1
                        logger.info(f"   Created: {referrer.get_full_name()}")
                    else:
                        stats['updated'] += 1
                        logger.info(f"   Updated: {referrer.get_full_name()}")

                except Exception as e:
                    logger.error(f"Error creating referrer '{coordinator_name}': {e}")
                    stats['errors'] += 1

        logger.info("")
        logger.info("=" * 70)
        logger.info("📊 Coordinator Extraction Summary")
        logger.info("=" * 70)
        logger.info(f"Unique coordinators found: {len(unique_coordinators):,}")
        logger.success(f"✅ Created: {stats['created']:,}")
        logger.info(f"🔄 Updated: {stats['updated']:,}")
        logger.info(f"⏭️  Skipped: {stats['skipped']:,}")
        logger.error(f"❌ Errors: {stats['errors']:,}")
        logger.info("")

    else:
        # Dry run - just show preview
        logger.info("")
        logger.info("=" * 70)
        logger.info("🔍 DRY RUN - Coordinator Preview")
        logger.info("=" * 70)
        logger.info(f"Would create {len(unique_coordinators)} coordinator records:")
        logger.info("")

        for i, (name, data) in enumerate(list(unique_coordinators.items())[:10], 1):
            logger.info(f"{i}. {name}")
            logger.info(f"   Phones: {', '.join(data['phones']) if data['phones'] else 'None'}")
            logger.info(f"   Emails: {', '.join(data['emails']) if data['emails'] else 'None'}")
            logger.info(f"   Notes: {len(data['notes'])} note(s)")
            logger.info("")

        if len(unique_coordinators) > 10:
            logger.info(f"... and {len(unique_coordinators) - 10} more coordinators")

        logger.info("")

    if dry_run or stats['errors'] == 0:
        logger.success("✅ ✅ ✅ Coordinator extraction completed successfully!")
        success = True
    else:
        logger.error("❌ Coordinator extraction completed with errors!")
        success = False

    logger.phase_end(success=success)
    return success


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Extract coordinators from Contacts.xlsx')
    parser.add_argument('--excel-file', type=str, help='Path to the Contacts.xlsx file')
    parser.add_argument('--dry-run', action='store_true', help='Preview only, do not save')
    args = parser.parse_args()

    success = extract_coordinators_from_contacts(excel_file=args.excel_file, dry_run=args.dry_run)
    sys.exit(0 if success else 1)

